# ============================================================
# CALM REPORTS — TEK HÜCRE COLAB KODU
# Bu kodu Colab'da tek bir hücreye yapıştır ve çalıştır.
# Müşterinin doldurduğu .xlsx dosyasını yükle, rapor hazır.
# ============================================================

# ---- KURULUM ------------------------------------------------
import subprocess, sys
subprocess.run([sys.executable, "-m", "pip", "install", "-q",
                "reportlab", "openpyxl"], check=True)

# ---- DOSYA YÜKLEME ------------------------------------------
from google.colab import files as colab_files
print("Müşteri xlsx dosyasını seçin...")
uploaded = colab_files.upload()
if not uploaded:
    raise SystemExit("Dosya seçilmedi.")
xlsx_path = list(uploaded.keys())[0]
print(f"Yüklendi: {xlsx_path}")

# ---- EMİSYON FAKTÖRLERİ ------------------------------------
FACTORS = {
    # malzemeler kg CO2e/kg
    "steel, primary": 2.45,
    "steel, recycled content": 1.55,
    "stainless steel": 4.20,
    "aluminium, primary": 12.50,
    "aluminium, recycled": 2.30,
    "copper": 3.50,
    "zinc": 3.30,
    "brass": 3.10,
    "glass": 1.20,
    "ceramic / porcelain": 0.75,
    "cement": 0.90,
    "concrete": 0.13,
    "gypsum / plasterboard": 0.39,
    "timber, softwood": 0.45,
    "timber, hardwood": 0.60,
    "plywood / mdf": 0.68,
    "paper / kraft board": 0.95,
    "cardboard, corrugated": 0.82,
    "eps (expanded polystyrene)": 3.20,
    "xps (extruded polystyrene)": 3.60,
    "pu foam / polyurethane": 4.10,
    "mineral wool": 1.35,
    "glass wool": 1.45,
    "pvc": 2.40,
    "hdpe": 2.00,
    "ldpe": 2.10,
    "polypropylene (pp)": 1.95,
    "pet": 2.70,
    "abs": 3.40,
    "nylon / polyamide": 6.50,
    "rubber, synthetic": 2.85,
    "adhesive, general": 2.60,
    "paint / coating": 2.90,
    "textile, cotton": 5.50,
    "textile, polyester": 4.30,
    "other (describe in notes)": 0.00,
    # enerji kg CO2e/birim
    "electricity (kwh)": 0.414,
    "natural gas (kwh)": 0.183,
    "natural gas (m3)": 2.020,
    "diesel (litres)": 2.680,
    "lpg (kg)": 2.940,
    "fuel oil (litres)": 3.180,
    "coal (kg)": 2.420,
    "district heat (kwh)": 0.170,
    # taşıma kg CO2e/t.km
    "road, hgv (>17t)": 0.1070,
    "road, hgv (7.5-17t)": 0.1850,
    "road, van (<3.5t)": 0.6000,
    "sea, container": 0.0159,
    "rail, freight": 0.0280,
    "air, freight": 0.6000,
    # ulaşım kg CO2e/km
    "passenger car, average": 0.1710,
}

def get_factor(name):
    key = name.strip().lower().replace(";", ",")
    if key in FACTORS:
        return FACTORS[key]
    for k, v in FACTORS.items():
        if key in k or k in key:
            return v
    raise KeyError(
        f"Faktör bulunamadı: '{name}'\n"
        f"Malzeme adı dropdown'dan seçilmeli veya FACTORS sözlüğüne eklenmeli.")

# ---- VERİ OKUMA ---------------------------------------------
from openpyxl import load_workbook

def num(v):
    if v is None or v == "":
        return None
    try:
        s = str(v).strip().replace(" ", "")
        if "," in s and "." in s:
            s = s.replace(",", "") if s.rfind(".") > s.rfind(",") \
                else s.replace(".", "").replace(",", ".")
        elif "," in s:
            s = s.replace(",", ".") if len(s.split(",")[-1]) != 3 \
                else s.replace(",", "")
        return float(s)
    except:
        return None

def txt(v):
    return "" if v is None else str(v).strip()

def find_sheet(wb, lead):
    for name in wb.sheetnames:
        if name.strip().startswith(str(lead)):
            return wb[name]
    return None

def read_workbook(path):
    wb = load_workbook(path, data_only=True)
    d = {"meta": {}, "products": [], "materials": [],
         "energy": [], "production": {}, "transport": [], "overhead": {}}

    sh = find_sheet(wb, 1)
    if sh:
        for row in sh.iter_rows(min_row=4, max_row=12, max_col=2, values_only=True):
            label = txt(row[0]).lower() if row[0] else ""
            val = row[1]
            if "company" in label or "şirket" in label:
                d["meta"]["company"] = txt(val)
            elif "facility" in label or "tesis" in label:
                d["meta"]["facility"] = txt(val)
            elif "start" in label or "başlangıç" in label:
                d["meta"]["period_start"] = txt(val)
            elif "end" in label or "bitiş" in label:
                d["meta"]["period_end"] = txt(val)

    sh = find_sheet(wb, 2)
    if sh:
        for row in sh.iter_rows(min_row=5, max_row=30,
                                max_col=4, values_only=True):
            code, name, wt, units = (list(row) + [None]*4)[:4]
            if not txt(code):
                continue
            if num(wt) and num(units):
                d["products"].append({"code": txt(code), "name": txt(name) or txt(code),
                                      "weight": num(wt), "units": num(units)})

    sh = find_sheet(wb, 3)
    if sh:
        for row in sh.iter_rows(min_row=5, max_row=120,
                                max_col=4, values_only=True):
            code, cat, desc, mass = (list(row) + [None]*4)[:4]
            if txt(code) and txt(cat) and num(mass):
                d["materials"].append({"code": txt(code), "category": txt(cat),
                                       "desc": txt(desc), "mass": num(mass)})

    sh = find_sheet(wb, 4)
    if sh:
        for row in sh.iter_rows(min_row=5, max_row=30,
                                max_col=15, values_only=True):
            row = list(row) + [None]*15
            src = txt(row[0])
            if not src:
                continue
            months = [num(row[i]) for i in range(1, 13)]
            provided = [v for v in months if v is not None]
            if not provided:
                continue
            d["energy"].append({"source": src, "amount": sum(provided),
                                 "months_present": len(provided),
                                 "note": txt(row[14])})

    sh = find_sheet(wb, 5)
    if sh:
        for row in sh.iter_rows(min_row=4, max_row=8,
                                max_col=2, values_only=True):
            label = txt(row[0]).lower() if row[0] else ""
            if ("total" in label or "toplam" in label) and num(row[1]):
                d["production"]["total_output_kg"] = num(row[1])
                break

    sh = find_sheet(wb, 6)
    if sh:
        for row in sh.iter_rows(min_row=5, max_row=60,
                                max_col=5, values_only=True):
            code, leg, mode, dist, mass = (list(row) + [None]*5)[:5]
            if txt(code) and txt(mode) and num(dist):
                d["transport"].append({"code": txt(code),
                                        "leg": txt(leg) or "Freight",
                                        "mode": txt(mode),
                                        "distance": num(dist),
                                        "mass": num(mass)})

    sh = find_sheet(wb, 7)
    if sh:
        ov = {}
        for row in sh.iter_rows(min_row=4, max_row=8,
                                max_col=2, values_only=True):
            label = txt(row[0]).lower() if row[0] else ""
            if "employee" in label or "çalışan" in label:
                ov["employees"] = num(row[1])
            elif "round" in label or "gidiş" in label:
                ov["round_trip_km"] = num(row[1])
            elif "working" in label or "çalışma" in label:
                ov["working_days"] = num(row[1])
        if all(ov.get(k) for k in ("employees", "round_trip_km", "working_days")):
            d["overhead"] = ov
    return d

# ---- HESAPLAMA ----------------------------------------------
def calculate(d):
    if not d["products"]:
        raise ValueError("Ürün bulunamadı — sekme 2'yi kontrol et.")
    total_output = d["production"].get("total_output_kg")
    if not total_output or total_output <= 0:
        raise ValueError("Toplam üretim eksik — sekme 5'i kontrol et.")

    prods = {p["code"]: dict(p, materials=[], transport=[])
             for p in d["products"]}

    for m in d["materials"]:
        if m["code"] not in prods:
            continue
        f = get_factor(m["category"])
        prods[m["code"]]["materials"].append(dict(m, factor=f,
                                                   emissions=m["mass"] * f))

    energy_lines, facility_total = [], 0.0
    for e in d["energy"]:
        if not e["amount"]:
            continue
        f = get_factor(e["source"])
        em = e["amount"] * f
        facility_total += em
        energy_lines.append(dict(e, factor=f, emissions=em))

    listed_mass = sum(p["weight"] * p["units"] for p in prods.values())
    if listed_mass > total_output * 1.001:
        raise ValueError(
            f"Ürün ağırlıkları ({listed_mass:,.0f} kg) toplam üretimi "
            f"({total_output:,.0f} kg) aşıyor. Sekme 5'teki rakamı kontrol et.")

    for p in prods.values():
        p["line_mass"] = p["weight"] * p["units"]
        p["share"] = p["line_mass"] / total_output
        p["energy_alloc"] = facility_total * p["share"]
        p["energy_per_unit"] = p["energy_alloc"] / p["units"]

    for t in d["transport"]:
        if t["code"] not in prods:
            continue
        p = prods[t["code"]]
        mass_kg = t["mass"] if t["mass"] else p["weight"]
        f = get_factor(t["mode"])
        tkm = mass_kg / 1000.0 * t["distance"]
        p["transport"].append(dict(t, mass_kg=mass_kg, factor=f,
                                   tkm=tkm, emissions=tkm * f))

    for p in prods.values():
        p["materials_per_unit"] = sum(m["emissions"] for m in p["materials"])
        p["transport_per_unit"] = sum(t["emissions"] for t in p["transport"])
        p["total_per_unit"] = (p["materials_per_unit"] + p["energy_per_unit"]
                               + p["transport_per_unit"])
        p["material_mass"] = sum(m["mass"] for m in p["materials"])

    overhead = None
    ov = d.get("overhead") or {}
    if all(ov.get(k) for k in ("employees", "round_trip_km", "working_days")):
        f = get_factor("Passenger car, average")
        km = ov["employees"] * ov["round_trip_km"] * ov["working_days"]
        total_c = km * f
        overhead = {"employees": ov["employees"],
                    "round_trip_km": ov["round_trip_km"],
                    "working_days": ov["working_days"],
                    "km": km, "factor": f, "total": total_c,
                    "per_product": {
                        c: {"alloc": total_c * p["share"],
                            "per_unit": total_c * p["share"] / p["units"]}
                        for c, p in prods.items()}}

    order = [p["code"] for p in d["products"]]
    warnings = []
    for c in order:
        p = prods[c]
        if not p["transport"]:
            warnings.append(f"{c}: taşıma verisi yok, hariç tutuldu.")
        gap = ((p["material_mass"] - p["weight"]) / p["weight"]
               if p["weight"] else 0)
        if abs(gap) > 0.15:
            warnings.append(
                f"{c}: malzeme toplamı {p['material_mass']:.3f} kg, "
                f"birim ağırlık {p['weight']:.3f} kg ({gap*100:+.0f}%).")
    for e in energy_lines:
        mp = e.get("months_present")
        if mp and mp < 12:
            warnings.append(
                f"Enerji '{e['source']}': 12 ayın {mp}'i dolu, "
                f"toplam o aylardan hesaplandı.")

    return {"meta": d["meta"], "order": order, "products": prods,
            "energy_lines": energy_lines, "facility_total": facility_total,
            "total_output": total_output, "overhead": overhead,
            "warnings": warnings}

# ---- PDF ÜRETİMİ -------------------------------------------
def build_pdf(res, out_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, PageBreak)
    from datetime import date

    PETROL  = colors.HexColor("#1E403B")
    SAGE_T  = colors.HexColor("#D7E2DB")
    PAPER2  = colors.HexColor("#E4EAE4")
    LINE    = colors.HexColor("#C3D0C8")
    INK     = colors.HexColor("#16211E")
    MUTED   = colors.HexColor("#57655F")
    HONEY_D = colors.HexColor("#B0812C")
    WHITE   = colors.white

    PW, PH = A4
    LM = RM = 20 * mm
    CW = PW - LM - RM

    def S(n, **kw):
        base = dict(fontName="Helvetica", fontSize=9.5, leading=14,
                    textColor=INK, alignment=TA_LEFT)
        base.update(kw)
        return ParagraphStyle(n, **base)

    st_h1   = S("h1", fontName="Helvetica-Bold", fontSize=18, leading=22, spaceAfter=6)
    st_h2   = S("h2", fontName="Helvetica-Bold", fontSize=13, leading=17,
                textColor=PETROL, spaceBefore=14, spaceAfter=6)
    st_body = S("b", spaceAfter=5)
    st_it   = S("i", fontName="Helvetica-Oblique", fontSize=8.5,
                leading=12, textColor=MUTED)
    st_hw   = S("hw", fontName="Helvetica-Bold", fontSize=8.5,
                leading=12, textColor=WHITE)
    st_hwr  = S("hwr", fontName="Helvetica-Bold", fontSize=8.5,
                leading=12, textColor=WHITE, alignment=TA_RIGHT)
    st_cell = S("c", fontSize=8.8, leading=12)
    st_cellb= S("cb", fontName="Helvetica-Bold", fontSize=8.8, leading=12)
    st_num  = S("n", fontName="Courier", fontSize=8.6, leading=12,
                alignment=TA_RIGHT, textColor=MUTED)
    st_numb = S("nb", fontName="Courier-Bold", fontSize=8.8, leading=12,
                alignment=TA_RIGHT, textColor=PETROL)

    meta  = res["meta"]
    order = res["order"]
    P     = res["products"]

    def tbl(data, widths, total_row=False):
        t = Table(data, colWidths=widths, repeatRows=1)
        cmds = [
            ("FONTNAME",     (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE",     (0, 0), (-1, -1), 8.8),
            ("VALIGN",       (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",   (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 5),
            ("LEFTPADDING",  (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
            ("BACKGROUND",   (0, 0), (-1,  0), PETROL),
            ("LINEBELOW",    (0, 0), (-1, -2), 0.35, LINE),
        ]
        for r in range(1, len(data)):
            if r % 2 == 0:
                cmds.append(("BACKGROUND", (0, r), (-1, r), PAPER2))
        if total_row:
            n = len(data) - 1
            cmds += [("BACKGROUND", (0, n), (-1, n), SAGE_T),
                     ("LINEABOVE",   (0, n), (-1, n), 0.8, PETROL)]
        t.setStyle(TableStyle(cmds))
        return t

    doc = SimpleDocTemplate(out_path, pagesize=A4,
                            leftMargin=LM, rightMargin=RM,
                            topMargin=22*mm, bottomMargin=20*mm)
    E = []

    # kapak bilgisi
    E.append(Paragraph("Product Carbon Footprint — Screening Estimate", st_h1))
    E.append(Spacer(1, 4))
    E.append(Paragraph(
        f"<b>Prepared for:</b> {meta.get('company','—')}<br/>"
        f"<b>Facility:</b> {meta.get('facility','—')}<br/>"
        f"<b>Period:</b> {meta.get('period_start','—')} – "
        f"{meta.get('period_end','—')}<br/>"
        f"<b>Products:</b> {len(order)}<br/>"
        f"<b>Issued:</b> {date.today().strftime('%d %B %Y')}<br/>"
        f"<b>Prepared by:</b> Calm Reports", st_body))
    E.append(Spacer(1, 6))
    E.append(Paragraph(
        "<i>Screening-level estimate — not a certified or verified LCA, "
        "and not an official CBAM declaration.</i>", st_it))
    E.append(Spacer(1, 14))

    # özet tablo
    E.append(Paragraph("Summary", st_h2))
    hdr = [Paragraph("Product", st_hw),
           Paragraph("Unit wt<br/>kg", st_hwr),
           Paragraph("Materials<br/>kg CO2e", st_hwr),
           Paragraph("Energy<br/>kg CO2e", st_hwr),
           Paragraph("Transport<br/>kg CO2e", st_hwr),
           Paragraph("<b>Total</b><br/>kg CO2e", st_hwr)]
    data = [hdr]
    for c in order:
        p = P[c]
        data.append([Paragraph(p["name"], st_cell),
                     Paragraph(f"{p['weight']:.3g}", st_num),
                     Paragraph(f"{p['materials_per_unit']:.2f}", st_num),
                     Paragraph(f"{p['energy_per_unit']:.2f}", st_num),
                     Paragraph(f"{p['transport_per_unit']:.2f}", st_num),
                     Paragraph(f"{p['total_per_unit']:.2f}", st_numb)])
    E.append(tbl(data, [52*mm, 18*mm, 24*mm, 20*mm, 24*mm, 22*mm]))
    E.append(Spacer(1, 8))

    # malzeme detayı
    E.append(Paragraph("Raw materials", st_h2))
    for c in order:
        p = P[c]
        if not p["materials"]:
            continue
        E.append(Paragraph(f"<b>{p['name']}</b>", st_cellb))
        E.append(Spacer(1, 3))
        hdr = [Paragraph("Material", st_hw),
               Paragraph("Mass<br/>kg/unit", st_hwr),
               Paragraph("Factor", st_hwr),
               Paragraph("Emissions<br/>kg CO2e", st_hwr)]
        data = [hdr]
        for m in p["materials"]:
            label = m["category"] + (f" — {m['desc']}" if m["desc"] else "")
            data.append([Paragraph(label, st_cell),
                         Paragraph(f"{m['mass']:.4g}", st_num),
                         Paragraph(f"{m['factor']:.3g}", st_num),
                         Paragraph(f"{m['emissions']:.3f}", st_numb)])
        data.append([Paragraph("<b>Total</b>", st_cellb),
                     Paragraph(f"<b>{p['material_mass']:.3f}</b>", st_numb),
                     Paragraph("", st_num),
                     Paragraph(f"<b>{p['materials_per_unit']:.3f}</b>", st_numb)])
        E.append(tbl(data, [68*mm, 22*mm, 22*mm, 28*mm], total_row=True))
        E.append(Spacer(1, 8))

    # tesis enerjisi
    if res["energy_lines"]:
        E.append(Paragraph("Facility energy", st_h2))
        hdr = [Paragraph("Source", st_hw),
               Paragraph("Annual total", st_hwr),
               Paragraph("Factor", st_hwr),
               Paragraph("Emissions<br/>kg CO2e", st_hwr),
               Paragraph("Months", st_hwr)]
        data = [hdr]
        for e in res["energy_lines"]:
            data.append([Paragraph(e["source"], st_cell),
                         Paragraph(f"{e['amount']:,.0f}", st_num),
                         Paragraph(f"{e['factor']:.4g}", st_num),
                         Paragraph(f"{e['emissions']:,.0f}", st_numb),
                         Paragraph(f"{e.get('months_present',12)}/12", st_num)])
        data.append([Paragraph("<b>Facility total</b>", st_cellb),
                     Paragraph("", st_num), Paragraph("", st_num),
                     Paragraph(f"<b>{res['facility_total']:,.0f}</b>", st_numb),
                     Paragraph("", st_num)])
        E.append(tbl(data, [50*mm, 28*mm, 22*mm, 28*mm, 18*mm], total_row=True))
        E.append(Spacer(1, 8))

        E.append(Paragraph("Allocation by mass", st_h2))
        E.append(Paragraph(
            f"Total output all products: {res['total_output']:,.0f} kg", st_body))
        hdr = [Paragraph("Product", st_hw),
               Paragraph("Units", st_hwr),
               Paragraph("Line mass<br/>kg", st_hwr),
               Paragraph("Share", st_hwr),
               Paragraph("Allocated<br/>kg CO2e", st_hwr),
               Paragraph("Per unit<br/>kg CO2e", st_hwr)]
        data = [hdr]
        for c in order:
            p = P[c]
            data.append([Paragraph(p["name"], st_cell),
                         Paragraph(f"{p['units']:,.0f}", st_num),
                         Paragraph(f"{p['line_mass']:,.0f}", st_num),
                         Paragraph(f"{p['share']*100:.1f}%", st_num),
                         Paragraph(f"{p['energy_alloc']:,.0f}", st_num),
                         Paragraph(f"{p['energy_per_unit']:.3f}", st_numb)])
        E.append(tbl(data, [45*mm, 18*mm, 22*mm, 18*mm, 26*mm, 22*mm]))
        E.append(Spacer(1, 8))

    # taşıma
    if any(P[c]["transport"] for c in order):
        E.append(Paragraph("Freight", st_h2))
        for c in order:
            p = P[c]
            if not p["transport"]:
                continue
            E.append(Paragraph(f"<b>{p['name']}</b>", st_cellb))
            E.append(Spacer(1, 3))
            hdr = [Paragraph("Leg", st_hw),
                   Paragraph("Mode", st_hwr),
                   Paragraph("km", st_hwr),
                   Paragraph("t.km", st_hwr),
                   Paragraph("kg CO2e", st_hwr)]
            data = [hdr]
            for t in p["transport"]:
                data.append([Paragraph(t["leg"], st_cell),
                             Paragraph(t["mode"], st_num),
                             Paragraph(f"{t['distance']:,.0f}", st_num),
                             Paragraph(f"{t['tkm']:.4f}", st_num),
                             Paragraph(f"{t['emissions']:.3f}", st_numb)])
            data.append([Paragraph("<b>Total freight</b>", st_cellb),
                         Paragraph("", st_num), Paragraph("", st_num),
                         Paragraph("", st_num),
                         Paragraph(f"<b>{p['transport_per_unit']:.3f}</b>",
                                   st_numb)])
            E.append(tbl(data, [42*mm, 38*mm, 18*mm, 22*mm, 22*mm],
                         total_row=True))
            E.append(Spacer(1, 8))

    # sonuçlar
    E.append(PageBreak())
    E.append(Paragraph("Results", st_h1))
    for c in order:
        p = P[c]
        E.append(Paragraph(f"<b>{p['name']}</b>", st_cellb))
        E.append(Spacer(1, 3))
        hdr = [Paragraph("Component", st_hw),
               Paragraph("kg CO2e / unit", st_hwr),
               Paragraph("Share", st_hwr)]
        data = [hdr]
        tot = p["total_per_unit"]
        for label, val in (("Raw materials", p["materials_per_unit"]),
                           ("Facility energy", p["energy_per_unit"]),
                           ("Freight", p["transport_per_unit"])):
            if val == 0 and label == "Freight":
                continue
            data.append([Paragraph(label, st_cell),
                         Paragraph(f"{val:.3f}", st_numb),
                         Paragraph(f"{val/tot*100:.0f}%" if tot else "—",
                                   st_num)])
        data.append([Paragraph("<b>Product footprint</b>", st_cellb),
                     Paragraph(f"<b>{tot:.3f}</b>", st_numb),
                     Paragraph("<b>100%</b>", st_num)])
        E.append(tbl(data, [72*mm, 42*mm, 36*mm], total_row=True))
        E.append(Spacer(1, 4))
        E.append(Paragraph(
            f"Reported as <b>{tot:.2f} kg CO2e per unit</b>, "
            f"cradle-to-gate.", st_body))
        E.append(Spacer(1, 10))

    # çalışan ulaşımı
    if res["overhead"]:
        ov = res["overhead"]
        E.append(Paragraph(
            "Corporate overhead — reported separately", st_h2))
        E.append(Paragraph(
            "Employee commuting is calculated but not added to the "
            "product footprint.", st_body))
        hdr = [Paragraph("Basis", st_hw), Paragraph("Value", st_hwr)]
        data = [hdr,
                [Paragraph("Employees", st_cell),
                 Paragraph(f"{ov['employees']:,.0f}", st_num)],
                [Paragraph("Round trip", st_cell),
                 Paragraph(f"{ov['round_trip_km']:,.0f} km", st_num)],
                [Paragraph("Working days", st_cell),
                 Paragraph(f"{ov['working_days']:,.0f}", st_num)],
                [Paragraph("Total distance", st_cell),
                 Paragraph(f"{ov['km']:,.0f} km", st_num)],
                [Paragraph("Commuting emissions", st_cell),
                 Paragraph(f"{ov['total']:,.0f} kg CO2e", st_num)]]
        for c in order:
            pu = ov["per_product"][c]["per_unit"]
            data.append([Paragraph(f"  → {P[c]['name']}", st_cell),
                         Paragraph(f"{pu:.3f} kg/unit", st_num)])
        E.append(tbl(data, [95*mm, 55*mm]))
        E.append(Spacer(1, 10))

    # sınırlama notu
    E.append(Paragraph("Statement of limitations", st_h2))
    E.append(Paragraph(
        "This is a screening-level estimate, not a certified Life Cycle "
        "Assessment, not verified by an accredited third party, and not "
        "an official CBAM compliance declaration. Where a verified figure "
        "is required, this report serves as structured input to that "
        "process.", st_body))

    # kullanılan faktörler
    E.append(Paragraph("Emission factors used", st_h2))
    used = set()
    for c in order:
        for m in P[c]["materials"]:
            used.add((m["category"], m["factor"], "kg CO2e/kg"))
        for t in P[c]["transport"]:
            used.add((t["mode"], t["factor"], "kg CO2e/t.km"))
    for e in res["energy_lines"]:
        unit = "kg CO2e/kWh" if "kwh" in e["source"].lower() else "kg CO2e/unit"
        used.add((e["source"], e["factor"], unit))
    if res["overhead"]:
        used.add(("Passenger car, average", res["overhead"]["factor"],
                  "kg CO2e/km"))
    hdr = [Paragraph("Input", st_hw),
           Paragraph("Factor", st_hwr),
           Paragraph("Unit", st_hw)]
    data = [hdr]
    for name, factor, unit in sorted(used):
        data.append([Paragraph(name, st_cell),
                     Paragraph(f"{factor:.4g}", st_numb),
                     Paragraph(unit, st_cell)])
    E.append(tbl(data, [65*mm, 28*mm, 57*mm]))

    doc.build(E)

# ---- ÇALIŞTIR -----------------------------------------------
print("\nVeri okunuyor...")
data = read_workbook(xlsx_path)

print("Hesap yapılıyor...")
res = calculate(data)

print("PDF üretiliyor...")
out_pdf = xlsx_path.replace(".xlsx", "_rapor.pdf").replace(".XLSX", "_rapor.pdf")
build_pdf(res, out_pdf)

print("\n" + "="*55)
print("SONUÇLAR")
print("="*55)
for c in res["order"]:
    p = res["products"][c]
    print(f"  {p['name'][:38]:40} {p['total_per_unit']:7.3f} kg CO2e/unit")
if res["overhead"]:
    lead = res["order"][0]
    pu = res["overhead"]["per_product"][lead]["per_unit"]
    print(f"  {'(+ çalışan ulaşımı, ayrı satır)':40} {pu:7.3f}")
if res["warnings"]:
    print("\nUYARILAR:")
    for w in res["warnings"]:
        print(f"  • {w}")
print(f"\nRapor: {out_pdf}")
colab_files.download(out_pdf)
